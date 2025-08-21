from datetime import datetime
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from firebase_admin import messaging
import requests
import os
import json
from .models import Users, Convocation
from rest_framework.response import Response
from rest_framework import status
from push_notifications.models import GCMDevice
import re
from requests.exceptions import RequestException
from typing import Optional, Dict
from openpyxl import load_workbook
import logging
logger = logging.getLogger(__name__)

def send_email_async(subject, message, from_email, recipient):
    try:
        email_message = EmailMultiAlternatives(
            subject=subject,
            body=message,
            from_email=f"No Reply <{from_email}>",
            to=[recipient],
        )
        email_message.attach_alternative(message, "text/html")
        email_message.send()
    except Exception as e:
        logger.error(f"Some error occured in send email, {e}")
    
def format_phone_number(phone_number: str) -> Optional[str]:
    """
    Format phone number to ensure it starts with +92 country code.

    Args:
        phone_number: Raw phone number string

    Returns:
        Formatted phone number with +92 prefix or None if invalid
    """
    if not phone_number:
        return None

    # Remove all non-digit characters
    cleaned = re.sub(r"[^\d]", "", phone_number)

    # Handle numbers that start with 0 (common in Pakistan)
    if cleaned.startswith("0"):
        return "+92" + cleaned[1:]

    # Handle numbers that start with 92 but missing +
    elif cleaned.startswith("92"):
        return "+" + cleaned

    # Handle numbers that already have +92
    elif cleaned.startswith("9292"):
        return "+" + cleaned[2:]

    # Handle numbers with no country code
    elif len(cleaned) == 10:  # Assuming 10-digit local numbers
        return "+92" + cleaned

    # Return None for invalid formats
    return None


def send_sms(user: Users, text: str) -> bool:
    """
    Send SMS to user's phone number using Veevotech API.

    Args:
        user: User object with phone_number attribute
        text: Message content to send

    Returns:
        bool: True if SMS was sent successfully, False otherwise
    """
    key = os.environ.get("SMS_KEY")
    if not key:
        print("SMS_KEY environment variable not set")
        return False

    # Get and validate phone number
    raw_number = getattr(user, "phone_number", None)
    if not raw_number:
        print("No phone number available for user")
        return False

    phone_number = format_phone_number(raw_number)
    if not phone_number:
        print(f"Invalid phone number format: {raw_number}")
        return False

    try:
        body = {
            "hash": key,
            "receivernum": phone_number,
            "textmessage": text,
            "sendernum": "Default",
        }

        headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        }

        response = requests.post(
            "https://api.veevotech.com/v3/sendsms",
            data=body,
            headers=headers,
            timeout=10,
        )

        response.raise_for_status()
        response_data = response.json()

        if response_data.get("STATUS") != "SUCCESSFUL":
            print(f"SMS API reported failure: {response_data}")
            return False

        return True

    except RequestException as e:
        print(f"Failed to send SMS: {str(e)}")
        if hasattr(e, "response") and e.response:
            print(f"API response: {e.response.text}")
        return False
    except Exception as e:
        print(f"Unexpected error sending SMS: {str(e)}")
        return False


def notify_user_devices(user: Users, title: str, body: str, url: str = None) -> bool:
    """
    Send a notification to a user's devices using Firebase Cloud Messaging.

    Args:
        user: The user to notify
        title: Notification title
        body: Notification body
        url: Optional URL to include in the notification data

    Returns:
        bool: True if notification was sent successfully, False otherwise
    """
    # Set default URL if none provided
    url = url or "/"

    # Get the FCM token - prioritize guest token if available
    fcm_token = user.guest_fcm_token

    # If no guest token, try to get registered devices
    if not fcm_token:
        print("No FCM token ")
        try:
            device = (
                GCMDevice.objects.filter(user=user).order_by("-date_created").first()
            )
            if not device:
                print("No FCM token available for user sending sms instead")
                return send_sms(user, body)
            fcm_token = device.registration_id  # Assuming this is the field name
        except Exception as e:
            print(f"Error getting user device: {str(e)}")
            return False

    # Create the FCM message
    message = messaging.Message(
        token=fcm_token,
        notification=messaging.Notification(
            title=title,
            body=body,
        ),
        data={
            "title": title,
            "body": body,
            "url": url,  # Include the URL in the data payload
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
        },
        android=messaging.AndroidConfig(
            priority="high",
            notification=messaging.AndroidNotification(
                channel_id="your_default_channel",
                click_action="FLUTTER_NOTIFICATION_CLICK",
                sound="default",
                icon="notification_icon",
            ),
        ),
        apns=messaging.APNSConfig(
            payload=messaging.APNSPayload(
                aps=messaging.Aps(
                    sound="default",
                    badge=1,
                ),
            ),
            headers={"apns-priority": "10"},  # Immediate delivery for iOS
        ),
    )

    try:
        response = messaging.send(message)
        print(f"Notification sent successfully: {response}")
        return True
    except Exception as e:
        print(f"Failed to send notification: {str(e)}")
        return False


def send_alert_email(
    to_email, subject, message, recipient_name="User", action_url=None
):
    context = {
        "subject": subject,
        "message": message,
        "recipient_name": recipient_name,
        "action_url": action_url,
        "site_name": "CustApp",
        "support_email": "support@custapp.pk",
    }

    html_content = render_to_string("email/alert_email.html", context)
    text_content = (
        f"{subject}\n\n{message}\n\nVisit: {action_url}"
        if action_url
        else f"{subject}\n\n{message}"
    )

    email = EmailMultiAlternatives(
        subject,
        text_content,
        f"No Reply <{settings.EMAIL_HOST_USER}>",  # <-- sender name and email
        [to_email]
    )
    email.attach_alternative(html_content, "text/html")
    email.connection = None
    email.send(fail_silently=False)


def add_comment_to_instance(request, instance, employee=None, student=None):
    try:
        text = request.data.get("text")
        if not text:
            return Response(
                {"error": "Comment text is required!"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        name = request.user.name
        user_type = request.user.user_type
        try:
            comments = (
                json.loads(instance.comments) if instance.comments else []
            )  # Load previous comment if exists
        except Exception:
            comments = []
        new_comment = {
            "name": name,
            "text": text,
            "user_type": user_type,
            "timestamp": datetime.now().isoformat(),
        }
        comments.append(new_comment)
        instance.comments = json.dumps(comments)
        instance.save()
        if employee and student:
            send_comment_notification(
                name=name,
                user_type=user_type,
                text=text,
                employee=employee,
                student=student,
            )
        return Response(
            {"message": "Saved sucessfully", "success": True},
            status=status.HTTP_201_CREATED,
        )
    except Exception as e:
        return Response(
            {"error": "Some error occured!", "details": str(e)},
            status=status.HTTP_400_BAD_REQUEST,
        )


def send_comment_notification(user_type, name, text, employee, student):
    #  = {
    #   payload  "head": "New Comment",
    #     "body": f"{name} commented: {text}",
    # }
    if user_type == "Student":
        # notify_user_devices(employee, "New comment on your applicaiton", body="You a new comment in your application!")
        send_alert_email(
            employee.email,
            "New Comment on Your Application",
            f"{name} commented: {text}",
            recipient_name=employee.name,
            action_url="/user/dashboard/",
        )
        notify_user_devices(
            employee, "New comment on your Application", f"{name} commented: {text}"
        )

    elif user_type == "Staff":
        send_alert_email(
            student.email,
            "New Comment on Your Application",
            f"{name} commented: {text}",
            recipient_name=student.name,
        )
        notify_user_devices(
            student, "New comment on your Application", f"{name} commented: {text}"
        )


def extract_year_term(reg_no: str) -> str:
    year_prefix = reg_no[:2]
    term_code = reg_no[2]
    term_map = {"1": "Spring", "3": "Fall"}
    year = f"20{year_prefix}"
    term = term_map.get(term_code, "Undefined Term Code")

    return f"{term} {year}"


def upload_student_data(
    file_path: str,
) -> None:

    total_updated = 0
    total_rows = 0
    not_exists = 0
    wb = load_workbook(filename=file_path, read_only=True)
    active = wb.active
    header = next(active.iter_rows(min_row=1, max_row=1, values_only=True))
    reg_idx = header.index("Reg #")
    father_name_idx = header.index("Father Name")
    email_idx = header.index("Personal Email")
    official_email = header.index("Official Email")
    cgpa_idx = header.index("CGPA")
    term_idx = header.index("Academic Term")
    is_active_idx = header.index("Academic Program/Active")

    for rows in active.iter_rows(min_row=2, values_only=True):  # Skip the header row
        reg_no = rows[reg_idx]
        is_active = rows[is_active_idx]
        if is_active:
            try:
                user = Users.objects.get(uu_id=reg_no)
                user.cgpa = rows[cgpa_idx]
                user.secondary_email = rows[email_idx]
                user.term = rows[term_idx]
                if not user.email:
                    user.email = rows[official_email]
                user.save()
                total_updated += 1
            except Users.DoesNotExist:
                print(f"The user with this {reg_no} does not exists")
                not_exists += 1
                continue

            total_rows += 1
        else:
            print(f"The user {reg_no} is not active!")

    print("-" * 10)
    print(f"Total rows  = {total_rows}")
    print(f"Total Users updated present in the DB {total_updated}")
    print(f"Total users to add = {not_exists}")


def load_convocation_data(file_path: str, convocation: Convocation) -> Dict:

    total_rows = 0
    users_modified = 0
    wb = load_workbook(filename=file_path, read_only=True)
    active = wb.active
    header = next(active.iter_rows(min_row=1, max_row=1, values_only=True))

    reg_idx = header.index("Registration No.")
    whatsapp_idx = header.index("Whatsapp Number")
    email_idx = header.index("Email Address")
    is_eligible_idx = header.index("Will you attend the Convocation?")
    for rows in active.iter_rows(min_row=2, values_only=True):  # Skip the header row
        reg_no = rows[reg_idx]

        try:
            user = Users.objects.get(uu_id=reg_no)
            # Update stuff
            if rows[is_eligible_idx].lower() == "yes":
                print("the user is eligible ")
                user.convocation = convocation
            if not user.secondary_email:
                print("Secondary email not found adding secondary email")
                user.secondary_email = rows[email_idx]
            user.phone_number = rows[whatsapp_idx]
            user.save()
            users_modified += 1
            print(f"{user.name} -- {user.phone_number}")
        except Users.DoesNotExist:
            print(f"The user with this {reg_no} does not exists")
            total_rows += 1
            continue
    return {"total": total_rows, "modified": users_modified}
