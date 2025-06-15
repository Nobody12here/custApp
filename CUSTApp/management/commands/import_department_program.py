import os
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from CUSTApp.models import (
    Department,
    Program,
)  # Replace 'your_app' with your actual app name


class Command(BaseCommand):
    help = "Import departments and programs from Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")

    def handle(self, *args, **options):
        file_path = options["file_path"]

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            # Create a dictionary to track created departments
            departments = {}

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                    short_code, dept_name, program_name = row

                    # Clean data (remove extra spaces, handle None values)
                    short_code = str(short_code).strip() if short_code else ""
                    dept_name = str(dept_name).strip() if dept_name else ""
                    program_name = str(program_name).strip() if program_name else ""

                    if not all([short_code, dept_name, program_name]):
                        self.stdout.write(
                            self.style.WARNING(f"Skipping incomplete row: {row}")
                        )
                        continue

                    # Get or create department
                    department, created = Department.objects.get_or_create(
                        dept_name=dept_name,
                        defaults={
                            "dept_name": dept_name,
                            # You'll need to set dept_head to an actual user or make it nullable
                            "dept_head_id": 6268,  # Temporary - set to an actual user ID
                        },
                    )

                    if created:
                        self.stdout.write(
                            self.style.SUCCESS(f"Created department: {dept_name}")
                        )
                        departments[dept_name] = department
                    else:
                        department = departments.get(dept_name, department)

                    # Create program
                    program, created = Program.objects.get_or_create(
                        program_name=program_name,
                        defaults={"program_name": program_name, "dept_id": department},
                    )

                    if created:
                        self.stdout.write(
                            self.style.SUCCESS(
                                f"Created program: {program_name} (Code: {short_code})"
                            )
                        )
                    else:
                        self.stdout.write(
                            self.style.WARNING(
                                f"Program already exists: {program_name}"
                            )
                        )

            self.stdout.write(self.style.SUCCESS("Import completed successfully!"))

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error during import: {str(e)}"))
