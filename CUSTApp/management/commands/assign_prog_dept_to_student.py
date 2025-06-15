import os
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from CUSTApp.models import Users, Program

class Command(BaseCommand):
    help = "Assign programs to students by matching program names from Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel-file",
            type=str,
            required=True,
            help="Path to Excel file with program mappings",
        )
        parser.add_argument(
            "--batch-size",
            type=int,
            default=1000,
            help="Number of students to process at a time",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_file"]
        batch_size = options["batch_size"]

        if not os.path.exists(excel_path):
            self.stderr.write(self.style.ERROR(f"Excel file not found: {excel_path}"))
            return

        try:
            # Step 1: Load program mappings from Excel
            self.stdout.write("Loading program mappings from Excel...")
            wb = load_workbook(filename=excel_path)
            sheet = wb.active

            # Create mapping from regno prefix to program name
            prefix_to_program_name = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                prefix, dept_name, program_name = row
                if prefix and program_name:
                    prefix_to_program_name[str(prefix).strip().upper()] = {
                        "program_name": str(program_name).strip(),
                        "dept_name": str(dept_name).strip(),
                    }

            self.stdout.write(f"Loaded {len(prefix_to_program_name)} mappings from Excel")

            # Step 2: Create mapping to database programs
            program_name_mapping = {}
            not_found_in_db = []

            for prefix, data in prefix_to_program_name.items():
                program_name = data["program_name"]
                try:
                    # Match by program name only (case-insensitive)
                    program = Program.objects.get(
                        program_name__iexact=program_name
                    )
                    program_name_mapping[prefix] = program
                except Program.DoesNotExist:
                    not_found_in_db.append(program_name)
                except Program.MultipleObjectsReturned:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Multiple programs found for name: {program_name}. Using first one."
                        )
                    )
                    program = Program.objects.filter(
                        program_name__iexact=program_name
                    ).first()
                    program_name_mapping[prefix] = program

            if not_found_in_db:
                self.stdout.write(
                    self.style.WARNING(
                        f"Could not find database programs for {len(not_found_in_db)} names: "
                        f"{', '.join(not_found_in_db)}"
                    )
                )

            # Step 3: Process students in batches
            total_students = Users.objects.count()
            processed = 0
            updated = 0
            no_match = 0
            no_regno = 0

            self.stdout.write(f"Processing {total_students} students...")

            while processed < total_students:
                with transaction.atomic():
                    students = Users.objects.filter(
                        user_type='Student'
                    )[processed : processed + batch_size]


                    for student in students:
                        processed += 1

                        if not student.uu_id:
                            no_regno += 1
                            continue

                        prefix = student.uu_id[:3].upper()
                        program = program_name_mapping.get(prefix)

                        if program:
                            Users.objects.filter(pk=student.pk).update(
                                program=program, 
                                dept=program.dept_id
                            )
                            updated += 1
                        else:
                            Users.objects.filter(pk=student.pk).update(
                                program=None, 
                                dept=None
                            )
                            no_match += 1

                    # Progress reporting
                    if processed % 1000 == 0 or processed == total_students:
                        self.stdout.write(
                            f"Processed {processed}/{total_students} "
                            f"({processed/total_students:.1%}) - "
                            f"Updated: {updated}, No match: {no_match}, No regno: {no_regno}"
                        )

            # Final report
            self.stdout.write(
                self.style.SUCCESS(
                    f"\nProcessing complete!\n"
                    f"Total students processed: {total_students}\n"
                    f"Successfully updated: {updated}\n"
                    f"No program match: {no_match}\n"
                    f"Missing registration numbers: {no_regno}\n"
                    f"Program names not found in database: {len(not_found_in_db)}"
                )
            )

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error during processing: {str(e)}"))