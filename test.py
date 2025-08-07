from openpyxl import load_workbook
from CUSTApp.models import Users, Convocation
from typing import Tuple
import os 

def upload_student_data(
    file_path: str,
) -> None:

    total_updated = 0
    total_rows = 0
    not_exists = 0
    wb = load_workbook(filename=file_path, read_only=True)
    active = wb.active
    header = next(active.iter_rows(min_row=1, max_row=1, values_only=True))
    reg_idx = header.index("uu_id")
    name_idx = header.index("Name")
    gender_idx = header.index("Gender")
    father_name_idx = header.index("Father Name")
    email_idx = header.index("Email personal")
    designation_idx = header.index("designation")

    for rows in active.iter_rows(min_row=2, values_only=True):  # Skip the header row
        reg_no = rows[reg_idx]
        name = rows[name_idx]
        father_name = rows[father_name_idx]
        email = rows[email_idx]
        gender = rows[gender_idx]
        designation = rows[designation_idx]
        user_type = "Staff"
        try:
            Users.objects.update_or_create(
            email=email,
            defaults={
                "uu_id": reg_no,
                "name": name,
                "father_name": father_name,
                "email": email,
                "user_type": user_type,
                "gender":gender,
                "designation":designation
            },
        )
            print("Updated or created sucessfully! for email = ",email)
        except Users.MultipleObjectsReturned:
            print("Multiple objects returned for this email ",email)
            continue
    print("-" * 10)
    print(f"Total rows  = {total_rows}")
    print(f"Total Users updated present in the DB {total_updated}")
    print(f"Total users to add = {not_exists}")

def upload_student_images(image_folder: str) -> None:
    
    try:
        images = os.listdir(image_folder)
        total_images = len(images)

        for index,image in enumerate(images,):
            image_path = os.path.join(image_folder, image)
            reg_no = image.split('.')[0]
            user = Users.objects.filter(uu_id=reg_no).first()
            if user:
                with open(image_path, 'rb') as img_file:
                    user.picture.save(image, img_file, save=True)
                percent = (index / total_images) * 100
                print(f"\rProcessing {index}/{total_images} images ({percent:.2f}%)", end='')

            else:
                continue

    except FileNotFoundError:
        print(f"Image folder {image_folder} does not exist.")
        return
upload_student_images('../../UniversityData/images')